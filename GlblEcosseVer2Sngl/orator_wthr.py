"""
#-------------------------------------------------------------------------------
# Name:
# Purpose:     read Excel spreadsheet of coordinates, retrieve weather and write ORATOR weather sheets
# Author:      Mike Martin
# Created:     11/12/2015
# Licence:     <your licence>
#-------------------------------------------------------------------------------
#
"""

__prog__ = 'orator_wthr.py'
__version__ = '0.0.1'
__author__ = 's03mm5'

from os.path import join, isfile, split
from os import remove
from openpyxl import load_workbook, Workbook
from zipfile import BadZipFile

import getClimGenNC
from getClimGenFns import associate_climate
import hwsd_bil

from string import ascii_uppercase
ALPHABET = list(ascii_uppercase)

sleepTime = 5
max_lines = 10
WARNING_MESS = '*** Warning *** '

def _write_block(strt_irow, wsht, nyears, year_list, met_list):
    '''
    proceed column major, row minor
    first row is a list of years
    '''
    indx = 0    # data pointer
    for iyr in range(nyears):
        irow = strt_irow
        col = ALPHABET[iyr]
        cell_loc = col + str(irow)
        wsht[cell_loc] = year_list[iyr]  # write year

        irow += 1
        for imnth in range(12):
            cell_loc = col + str(irow)
            wsht[cell_loc] = met_list[indx]
            indx += 1
            irow += 1

    return irow + 1

def generate_orator_wthr(form):
    '''
    extract coordinates and retrieve weather
    '''
    nyears = 10
    year_end = form.weather_sets['CRU_hist']['year_end']
    year_start = year_end - nyears + 1
    year_list = []
    for yr in range(year_start, year_end + 1):
        year_list.append(yr)

    coords_xls = form.w_lbl16.text()
    coord_list = _fetch_coordinates(coords_xls)
    out_dir, short_fname = split(coords_xls)

    # retrieve weather from datasets
    # ==============================

    snglPntFlag = True
    num_band = 0
    hwsd = hwsd_bil.HWSD_bil(form.lgr, form.hwsd_dir)
    climgen = getClimGenNC.ClimGenNC(form)

    for area_name, bbox_aoi, lat, lon in coord_list:

        # generate weather dataset indices which enclose the AOI
        # ======================================================
        aoi_indices_fut, aoi_indices_hist = climgen.genLocalGrid(bbox_aoi, hwsd, snglPntFlag, num_band)

        pettmp_fut = climgen.fetch_cru_future_NC_data(aoi_indices_fut, num_band)
        pettmp_hist = climgen.fetch_cru_historic_NC_data(aoi_indices_hist, num_band)

        gran_lat = int(round((90.0 - lat) * hwsd.granularity))
        gran_lon = int(round((180.0 + lon) * hwsd.granularity))

        site_rec = list([gran_lat, gran_lon, lat, lon, None, None])
        pettmp_grid_cell = associate_climate(site_rec, climgen, pettmp_hist, pettmp_fut)
        nvals = len(pettmp_grid_cell['precipitation'][0])

        # only interested in historic data
        # ================================
        pettmp = {}
        for metric in pettmp_grid_cell:
            indx = nvals - 12*nyears      # ten years
            pettmp[metric] = pettmp_grid_cell[metric][0][indx:]

        print('{} {} {} {}'.format(area_name, lat, lon, len(pettmp[metric])))

        # construct output
        # ================
        fname_out = join(out_dir, area_name + '.xlsx')
        if isfile(fname_out):
            remove(fname_out)

        wb_obj = Workbook()
        wsht = wb_obj.create_sheet('Weather')
        del wb_obj['Sheet']     # created by default


        irow = 1
        for period in ['steady state', 'forward run']:
            for metric in pettmp:
                met_list = pettmp[metric]
                irow = _write_block(irow, wsht, nyears, year_list, met_list)

        wb_obj.save(fname_out)
        print('wrote ' + fname_out)

    return

def _fetch_coordinates(coords_xls):
    '''
    extract coordinates from Excel file
    '''
    try:
        wb_obj = load_workbook(coords_xls, data_only=True)
        sheet_name = wb_obj.sheetnames[0]
    except (PermissionError, BadZipFile) as err:
        print('Error: ' + str(err))
        return None

    # gather data for each area by deconstructing each cell
    # =====================================================
    sheet = wb_obj[sheet_name]
    coord_list = []
    for irow in range(2, sheet.max_row):
        val = sheet.cell(row=irow, column=1).value
        if val is None:
            continue
        elif val.find('Area') >= 0:
            area_name = val[5:]

            # read four coordinates
            # =====================
            lat_min, lon_min = 2 * [99999]
            lat_max, lon_max = 2 * [-99999]
            indx1 = irow + 1
            indx2 = irow + 4
            for irow in range(indx1, indx2):
                val = sheet.cell(row = irow, column = 1).value
                if val is not None:
                    lat, lon = [sheet.cell(row=irow, column=2).value, sheet.cell(row=irow, column=3).value]
                    lat_max = max(lat_max, lat)
                    lat_min = min(lat_min, lat)

                    lon_max = max(lon_max, lon)
                    lon_min = min(lon_min, lon)

            lat = (lat_max + lat_min)/2
            lon = (lon_max + lon_min)/2
            bbox_aoi = list([lon - 0.01, lat - 0.01, lon + 0.01, lat + 0.01])
            coord_list.append([area_name, bbox_aoi, lat, lon])

    wb_obj.close()

    return coord_list
